from openpyxl import Workbook
import os
import json
from time import sleep

# Explanation:
# This script is used to get the record of a Luogu user and save it to an Excel file.
# The script uses the GetRecordPage.bat file to get the data of each page and then formats the data using FormatJson.py.

# Warning:
# This script may be blocked by Luogu if you request too many pages at once.
# Do not set the sleep_time too low or Luogu may block your IP.

# Warning:
# This script would not get the record pages if the old record files are not deleted.
# You have to delete the old record files manually before running the script.

# Warning:
# This script may not work if the Luogu website has changed.

# You can make a issue on GitHub if you find any problem.

# 配置信息
sleep_time = 1 # 获取间隔时间(秒)，防止请求过快被封IP
uid = "" # 你的UID
client_id = "" # 你的client_id(cookie)
range_max = 1 # 最大页数


wb = Workbook()
ws = wb.active
cnt = 1

# 设置表头
ws[f"B{1}"]= "题目ID"
ws[f"C{1}"]= "题目名称"
ws[f"D{1}"]= "题目难度"
ws[f"E{1}"]= "提交时间"
ws[f"F{1}"]= "提交分数"
ws[f"G{1}"]= "消耗时间"
ws[f"H{1}"]= "内存占用"
ws[f"I{1}"]= "代码长度"
ws[f"J{1}"]= "提交状态"

# 循环获取数据并写入Excel
for i in range(1, range_max+1):
    print(f"正在获取第{i}页数据...")
    # 如果文件不存在则调用GetRecordPage.bat获取数据并格式化
    if not os.path.exists(f"./RecordFiles/record{i}_Formatted.json"):
        os.system(f"GetRecordPage.bat {uid} {client_id} {i} > log.txt")
        os.system(f"python FormatJson.py ./RecordFiles/record{i}.json ./RecordFiles/record{i}_Formatted.json")
    # 读取格式化后的json文件
    with open(f"./RecordFiles/record{i}_Formatted.json", 'r', encoding='utf-8') as file:
        in_data = json.load(file)
    in_data = in_data['currentData']['records']['result']
    for plm in in_data:
        cnt += 1
        ws[f"A{cnt}"] = cnt-1
        ws[f"B{cnt}"] = plm['problem']['pid']
        ws[f"C{cnt}"] = plm['problem']['title']
        ws[f"D{cnt}"] = plm['problem']['difficulty']
        ws[f"E{cnt}"] = plm['submitTime']
        if 'score' in plm:
            ws[f"F{cnt}"] = plm['score']
        ws[f"G{cnt}"] = plm['time']
        ws[f"H{cnt}"] = plm['memory']
        ws[f"I{cnt}"] = plm['sourceCodeLength']
        ws[f"J{cnt}"] = plm['status']
    # 清除旧的json文件
    if os.path.exists(f"./RecordFiles/record{i}.json"):
        os.remove(f"./RecordFiles/record{i}.json")
        sleep(1)
    if i < range_max:
        os.system(f"cls")

wb.save("LuoguRecord.xlsx")
